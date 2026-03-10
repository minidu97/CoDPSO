import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side)
from openpyxl.utils import get_column_letter
import numpy as np

# ----------------------------------------------------------------
# Paper's original results from Tables 15-18
# Format: {fid: (mean, std, rank)}
# ----------------------------------------------------------------
PAPER_RESULTS_D10 = {
    1:  (300.00,    0.00e+00),
    2:  (400.78,    1.95e+00),
    3:  (600.00,    8.98e-03),
    4:  (806.83,    2.67e+00),
    5:  (900.01,    4.13e-02),
    6:  (1871.86,   5.91e+01),
    7:  (2009.35,   9.66e+00),
    8:  (2210.06,   1.00e+01),
    9:  (2529.28,   0.00e+00),
    10: (2500.64,   5.18e-01),
    11: (2610.00,   5.48e+01),
    12: (2861.61,   9.46e-01),
}

PAPER_RESULTS_D20 = {
    1:  (300.00,    1.83e-14),
    2:  (422.16,    2.88e+01),
    3:  (600.53,    7.77e-01),
    4:  (831.46,    6.29e+00),
    5:  (905.96,    6.80e+00),
    6:  (2050.76,   2.10e+02),
    7:  (2035.73,   1.37e+01),
    8:  (2223.40,   3.44e+00),
    9:  (2480.78,   1.17e-12),
    10: (2516.76,   5.38e+01),
    11: (2853.42,   1.58e+02),
    12: (2947.63,   3.88e+00),
}

FUNC_NAMES = {
    1:  "F1 - Zakharov",
    2:  "F2 - Rosenbrock",
    3:  "F3 - Schaffer f6",
    4:  "F4 - Rastrigin",
    5:  "F5 - Levy",
    6:  "F6 - Hybrid 1",
    7:  "F7 - Hybrid 2",
    8:  "F8 - Hybrid 3",
    9:  "F9 - Composition 1",
    10: "F10 - Composition 2",
    11: "F11 - Composition 3",
    12: "F12 - Composition 4",
}

# ----------------------------------------------------------------
# Style helpers
# ----------------------------------------------------------------
HEADER_FILL    = PatternFill("solid", start_color="2F4F8F", end_color="2F4F8F")
SUBHEAD_FILL   = PatternFill("solid", start_color="4F81BD", end_color="4F81BD")
PAPER_FILL     = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
MY_FILL        = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
DIFF_FILL      = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
BETTER_FILL    = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
WORSE_FILL     = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")

WHITE_FONT     = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BOLD_FONT      = Font(name="Arial", bold=True, size=10)
NORMAL_FONT    = Font(name="Arial", size=10)
CENTER         = Alignment(horizontal="center", vertical="center")
LEFT           = Alignment(horizontal="left",   vertical="center")

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def style_cell(cell, font=None, fill=None, alignment=None, border=True):
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = thin_border()


def write_results_sheet(wb, sheet_name, my_results, dim):
    """
    Write one sheet comparing paper results vs my results.

    my_results : dict {fid: np.ndarray of shape (30,)}
    dim        : int  10 or 20
    """
    ws = wb.create_sheet(title=sheet_name)
    paper_results = PAPER_RESULTS_D10 if dim == 10 else PAPER_RESULTS_D20

    # ---- Title row ----
    ws.merge_cells("A1:K1")
    ws["A1"] = f"CoDPSO Results Comparison — CEC 2022 (D={dim})"
    style_cell(ws["A1"], font=WHITE_FONT, fill=HEADER_FILL,
               alignment=CENTER, border=False)
    ws.row_dimensions[1].height = 28

    # ---- Column headers ----
    headers = [
        "Function",
        "Paper Mean", "Paper Std",
        "My Mean",    "My Std",    "My Best",  "My Worst",
        "Diff (Mean)", "Diff (Std)",
        "Mean Better?", "Notes"
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        style_cell(cell, font=WHITE_FONT, fill=SUBHEAD_FILL, alignment=CENTER)
    ws.row_dimensions[2].height = 22

    # ---- Data rows ----
    for row_idx, fid in enumerate(range(1, 13), start=3):
        p_mean, p_std = paper_results[fid]

        data    = np.array(my_results.get(fid, [np.nan] * 30))
        m_mean  = float(np.mean(data))
        m_std   = float(np.std(data))
        m_best  = float(np.min(data))
        m_worst = float(np.max(data))

        diff_mean = m_mean - p_mean
        diff_std  = m_std  - p_std
        better    = "✅ Better" if m_mean < p_mean else ("➖ Equal" if m_mean == p_mean else "❌ Worse")

        row_fill = MY_FILL if row_idx % 2 == 0 else PAPER_FILL

        values = [
            FUNC_NAMES[fid],
            p_mean, p_std,
            m_mean, m_std, m_best, m_worst,
            diff_mean, diff_std,
            better, ""
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            style_cell(cell, font=NORMAL_FONT, fill=row_fill, alignment=CENTER)

            # Highlight diff column
            if col == 8:
                cell.fill = BETTER_FILL if diff_mean < 0 else (
                    WORSE_FILL if diff_mean > 0 else row_fill
                )
            # Highlight better/worse column
            if col == 10:
                cell.fill = (BETTER_FILL if "Better" in str(val)
                             else WORSE_FILL if "Worse" in str(val)
                             else row_fill)

        # Number format for scientific notation columns
        for col in [2, 3, 4, 5, 6, 7, 8, 9]:
            ws.cell(row=row_idx, column=col).number_format = "0.00E+00"

    # ---- Summary row ----
    summary_row = 15
    ws.merge_cells(f"A{summary_row}:C{summary_row}")
    ws[f"A{summary_row}"] = "Summary"
    style_cell(ws[f"A{summary_row}"], font=WHITE_FONT,
               fill=HEADER_FILL, alignment=CENTER)

    ws[f"D{summary_row}"] = f'=COUNTIF(J3:J14,"*Better*")'
    ws[f"E{summary_row}"] = "Functions Better"
    ws[f"F{summary_row}"] = f'=COUNTIF(J3:J14,"*Worse*")'
    ws[f"G{summary_row}"] = "Functions Worse"
    ws[f"H{summary_row}"] = f'=COUNTIF(J3:J14,"*Equal*")'
    ws[f"I{summary_row}"] = "Functions Equal"

    for col in [4, 5, 6, 7, 8, 9]:
        style_cell(ws.cell(summary_row, col),
                   font=BOLD_FONT, fill=DIFF_FILL, alignment=CENTER)

    # ---- Column widths ----
    col_widths = [22, 14, 14, 14, 14, 14, 14, 16, 14, 14, 18]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def write_raw_data_sheet(wb, my_results, dim):
    """
    Write a sheet with all 30 raw run values per function.
    """
    ws = wb.create_sheet(title=f"Raw Data D={dim}")

    # Header
    ws.merge_cells("A1:AM1")
    ws["A1"] = f"Raw Run Results — CEC 2022 (D={dim}) — 30 Independent Runs"
    style_cell(ws["A1"], font=WHITE_FONT, fill=HEADER_FILL,
               alignment=CENTER, border=False)
    ws.row_dimensions[1].height = 24

    # Column headers: Function | Run 1 | Run 2 | ... | Run 30 | Mean | Std | Best
    ws.cell(2, 1, "Function")
    style_cell(ws.cell(2, 1), font=WHITE_FONT,
               fill=SUBHEAD_FILL, alignment=CENTER)

    for r in range(1, 31):
        cell = ws.cell(2, r + 1, f"Run {r}")
        style_cell(cell, font=WHITE_FONT, fill=SUBHEAD_FILL, alignment=CENTER)

    for col, label in zip([32, 33, 34, 35], ["Mean", "Std", "Best", "Worst"]):
        cell = ws.cell(2, col, label)
        style_cell(cell, font=WHITE_FONT, fill=SUBHEAD_FILL, alignment=CENTER)

    ws.row_dimensions[2].height = 20

    # Data
    for row_idx, fid in enumerate(range(1, 13), start=3):
        data = np.array(my_results.get(fid, [np.nan] * 30))
        row_fill = MY_FILL if row_idx % 2 == 0 else PAPER_FILL

        ws.cell(row_idx, 1, FUNC_NAMES[fid])
        style_cell(ws.cell(row_idx, 1),
                   font=BOLD_FONT, fill=row_fill, alignment=LEFT)

        for r_idx, val in enumerate(data[:30], start=2):
            cell = ws.cell(row_idx, r_idx, float(val))
            cell.number_format = "0.00E+00"
            style_cell(cell, font=NORMAL_FONT, fill=row_fill, alignment=CENTER)

        # Summary formulas
        start_col = get_column_letter(2)
        end_col   = get_column_letter(31)
        rng       = f"{start_col}{row_idx}:{end_col}{row_idx}"

        for col, formula in zip(
            [32, 33, 34, 35],
            [f"=AVERAGE({rng})", f"=STDEV({rng})",
             f"=MIN({rng})",     f"=MAX({rng})"]
        ):
            cell = ws.cell(row_idx, col, formula)
            cell.number_format = "0.00E+00"
            style_cell(cell, font=BOLD_FONT, fill=DIFF_FILL, alignment=CENTER)

    # Column widths
    ws.column_dimensions["A"].width = 22
    for col in range(2, 36):
        ws.column_dimensions[get_column_letter(col)].width = 12


def export_results_to_excel(my_results_d10, my_results_d20, filepath="CoDPSO_Results.xlsx"):
    """
    Main export function. Call this from main.py after experiments.

    Parameters
    ----------
    my_results_d10 : dict {fid: list/array of 30 run results}  for D=10
    my_results_d20 : dict {fid: list/array of 30 run results}  for D=20
    filepath       : str  output file path
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    # Sheet 1: D=10 comparison
    write_results_sheet(wb, "D=10 Comparison", my_results_d10, dim=10)

    # Sheet 2: D=20 comparison
    write_results_sheet(wb, "D=20 Comparison", my_results_d20, dim=20)

    # Sheet 3: Raw data D=10
    write_raw_data_sheet(wb, my_results_d10, dim=10)

    # Sheet 4: Raw data D=20
    write_raw_data_sheet(wb, my_results_d20, dim=20)

    wb.save(filepath)
    print(f"Results saved to: {filepath}")